"""
Migración de ventas históricas desde:
  Copia de Planilla del dia FINAL.xlsx  →  Neon PostgreSQL

Uso:
  DATABASE_URL="postgresql://..." python3 scripts/migrate_ventas_historicas.py

Estructura del Excel (por hoja):
  Col A = Fecha | B = Día | C = INDEP Total | D = vacía
  E = INDEP Tarjetas | F = INDEP Tickets
  G = LURO Total   | H = vacía
  I = LURO Tarjetas | J = LURO Tickets
"""

import os, sys
import openpyxl
from datetime import datetime, date

# ── Conexión ──────────────────────────────────────────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "")
if not DATABASE_URL:
    print("❌  Falta DATABASE_URL. Ejecutá:")
    print('   DATABASE_URL="postgresql://..." python3 scripts/migrate_ventas_historicas.py')
    sys.exit(1)

try:
    import psycopg2
except ImportError:
    print("Instalando psycopg2...")
    os.system("pip3 install psycopg2-binary -q")
    import psycopg2

# Corregir prefijo si es necesario (Render a veces usa postgres://)
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

EXCEL_PATH = "/Users/juanavila/Library/CloudStorage/OneDrive-Personal/SUR MADERAS/Administrativo/Sucursales 2025/Copia de Planilla del dia FINAL.xlsx"

MONTHS_ES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

# Hojas a procesar: (nombre_hoja, mes_label, year)
SHEETS_TO_PROCESS = [
    ("ENERO 2025",     "ENERO",     2025),
    ("FEBRERO 2025",   "FEBRERO",   2025),
    ("MARZO 2025",     "MARZO",     2025),
    ("ABRIL 2025",     "ABRIL",     2025),
    ("MAYO 2025",      "MAYO",      2025),
    ("JUNIO 2025",     "JUNIO",     2025),
    ("JULIO 2025",     "JULIO",     2025),
    ("AGOSTO 2025",    "AGOSTO",    2025),
    ("SEPTIEMBRE",     "SEPTIEMBRE",2025),
    ("OCTUBRE 2025",   "OCTUBRE",   2025),
    ("NOVIEMBRE 2025", "NOVIEMBRE", 2025),
    ("DICIEMBRE 2025", "DICIEMBRE", 2025),
    ("ENERO 2026",     "ENERO",     2026),
    ("FEBRERO 2026",   "FEBRERO",   2026),
    ("MARZO 2026",     "MARZO",     2026),
    ("ABRIL 2026",     "ABRIL",     2026),
    ("MAYO 2026",      "MAYO",      2026),
]

# branch_id: 2 = INDEPENDENCIA, 1 = LURO
BRANCH_INDEP = 2
BRANCH_LURO  = 1

def safe_float(val):
    try:
        if val is None: return None
        return float(val)
    except: return None

def safe_int(val):
    try:
        if val is None: return None
        return int(float(val))
    except: return None

def parse_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    return None


def migrate():
    print(f"\n🔌  Conectando a Neon...")
    conn = psycopg2.connect(DATABASE_URL, sslmode='require')
    cur  = conn.cursor()
    print("✅  Conectado.\n")

    # Cargar Excel
    print("📂  Abriendo Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"✅  Hojas disponibles: {wb.sheetnames}\n")

    total_inserted = 0
    total_updated  = 0
    total_skipped  = 0

    for sheet_name, month_label, year in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️   Hoja '{sheet_name}' no encontrada — saltando.")
            continue

        ws = wb[sheet_name]
        print(f"📋  Procesando {sheet_name} ({month_label} {year})...")

        rows_this_sheet = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            sale_date = parse_date(row[0])
            if sale_date is None:
                continue  # fila de total o vacía

            # Skip si es domingo (el negocio está cerrado)
            if sale_date.weekday() == 6:
                continue

            # Verificar que la fecha corresponde al mes/año correcto
            if sale_date.year != year or sale_date.month != MONTHS_ES.get(month_label, 0):
                continue

            indep_total  = safe_float(row[2])
            indep_cards  = safe_float(row[4])
            indep_tickets = safe_int(row[5])

            luro_total   = safe_float(row[6])
            luro_cards   = safe_float(row[8])
            luro_tickets  = safe_int(row[9])

            # Insertar/actualizar cada sucursal
            for branch_id, total, cards, tickets in [
                (BRANCH_INDEP, indep_total, indep_cards, indep_tickets),
                (BRANCH_LURO,  luro_total,  luro_cards,  luro_tickets),
            ]:
                if total is None and cards is None and tickets is None:
                    total_skipped += 1
                    continue  # sin datos para este día/sucursal

                # ¿Ya existe?
                cur.execute(
                    "SELECT id FROM daily_sales WHERE sale_date = %s AND branch_id = %s",
                    (sale_date, branch_id)
                )
                existing = cur.fetchone()

                if existing:
                    cur.execute("""
                        UPDATE daily_sales
                        SET total_amount = %s, card_payments = %s, ticket_count = %s,
                            month_label = %s, year = %s
                        WHERE id = %s
                    """, (total, cards, tickets, month_label, year, existing[0]))
                    total_updated += 1
                else:
                    cur.execute("""
                        INSERT INTO daily_sales
                            (sale_date, branch_id, total_amount, card_payments,
                             ticket_count, month_label, year, closed)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (sale_date, branch_id, total, cards, tickets, month_label, year, True))
                    total_inserted += 1

                rows_this_sheet += 1

        conn.commit()
        print(f"   ✅  {rows_this_sheet} registros procesados para {sheet_name}")

    cur.close()
    conn.close()

    print(f"""
╔══════════════════════════════════════╗
║         MIGRACIÓN COMPLETA           ║
╠══════════════════════════════════════╣
║  ✅ Insertados:  {total_inserted:<20} ║
║  🔄 Actualizados:{total_updated:<20} ║
║  ⏭️  Saltados:   {total_skipped:<20} ║
╚══════════════════════════════════════╝
""")

if __name__ == "__main__":
    migrate()
