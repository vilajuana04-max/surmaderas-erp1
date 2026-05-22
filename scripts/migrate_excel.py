"""
Script de migración: SurMaderas_ERP.xlsm → PostgreSQL

Migra:
  - _COMPRAS_HIST  → purchases + providers   (1.578 registros históricos)
  - _REGISTRO_VENTAS → daily_sales           (ventas históricas cerradas)
  - Vacaciones 2024/2025/2026 → vacation_records
  - EMPLEADOS → employees (actualiza fechas de ingreso)

Uso:
  pip install openpyxl psycopg2-binary
  python scripts/migrate_excel.py
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values
from datetime import datetime, date

EXCEL_PATH = "/Users/juanavila/Library/CloudStorage/OneDrive-Personal/SUR MADERAS/Administrativo/SurMaderas_ERP.xlsm"
DB_URL     = os.getenv("DATABASE_URL", "postgresql://surmaderas:surmaderas2026@localhost:5432/surmaderas")

MONTHS_MAP = {
    "Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
    "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,"Noviembre":11,"Diciembre":12,
    "ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,
    "JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12,
}

MONTHS_LABEL = {
    1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
    7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE",
}


def to_date(v):
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    return None


def to_float(v):
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def migrate():
    print("📖 Leyendo Excel...")
    wb = load_workbook(EXCEL_PATH, data_only=True, keep_vba=True)
    conn = psycopg2.connect(DB_URL)
    cur  = conn.cursor()

    # ─── 1. Proveedores + Compras históricas ─────────────────────
    print("🛒 Migrando _COMPRAS_HIST...")
    ws = wb["_COMPRAS_HIST"]
    providers_cache = {}
    purchases_rows  = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        mes_registro, fecha, nro_factura, proveedor, total, flag = row[:6]
        if not proveedor or total is None:
            continue

        proveedor = str(proveedor).strip().upper()
        if proveedor not in providers_cache:
            cur.execute(
                "INSERT INTO providers (name) VALUES (%s) ON CONFLICT (name) DO UPDATE SET name=EXCLUDED.name RETURNING id",
                (proveedor,)
            )
            providers_cache[proveedor] = cur.fetchone()[0]

        purchase_date = to_date(fecha)
        total_amount  = to_float(total)
        if total_amount is None:
            continue

        month_label = None
        year        = None
        if isinstance(mes_registro, str):
            parts = mes_registro.strip().split()
            if len(parts) >= 2:
                month_label = parts[0].upper()
                try:
                    year = int(parts[1])
                except ValueError:
                    pass
        if not month_label and purchase_date:
            month_label = MONTHS_LABEL[purchase_date.month]
            year        = purchase_date.year

        purchases_rows.append((
            purchase_date,
            str(nro_factura).strip() if nro_factura else None,
            providers_cache[proveedor],
            total_amount,
            str(flag).strip() if flag else None,
            month_label,
            year,
            True,  # closed = True (históricas)
        ))

    if purchases_rows:
        execute_values(cur, """
            INSERT INTO purchases
              (purchase_date, invoice_number, provider_id, total_amount, flag, month_label, year, closed)
            VALUES %s
            ON CONFLICT DO NOTHING
        """, purchases_rows)
        print(f"  ✅ {len(purchases_rows)} facturas migradas")

    # ─── 2. Ventas históricas ─────────────────────────────────────
    print("💰 Migrando _REGISTRO_VENTAS...")
    ws_v = wb["_REGISTRO_VENTAS"]
    sales_rows = []

    # Branch IDs: LURO=1, INDEPENDENCIA=2
    for row in ws_v.iter_rows(min_row=3, values_only=True):
        mes, anio, fecha, dia, indep_total, indep_tarjetas, indep_tickets, indep_tprom, \
        luro_total, luro_tarjetas, luro_tickets, luro_tprom, total_dia = (row + (None,)*13)[:13]

        sale_date = to_date(fecha)
        if not sale_date:
            continue

        month_label = str(mes).upper() if mes else MONTHS_LABEL.get(sale_date.month)
        year        = int(anio) if anio else sale_date.year

        if to_float(indep_total) is not None:
            sales_rows.append((sale_date, 2, to_float(indep_total),
                               to_float(indep_tarjetas), int(indep_tickets) if indep_tickets else None,
                               month_label, year, True))
        if to_float(luro_total) is not None:
            sales_rows.append((sale_date, 1, to_float(luro_total),
                               to_float(luro_tarjetas), int(luro_tickets) if luro_tickets else None,
                               month_label, year, True))

    if sales_rows:
        execute_values(cur, """
            INSERT INTO daily_sales
              (sale_date, branch_id, total_amount, card_payments, ticket_count, month_label, year, closed)
            VALUES %s
            ON CONFLICT (sale_date, branch_id) DO NOTHING
        """, sales_rows)
        print(f"  ✅ {len(sales_rows)} registros de ventas migrados")

    # ─── 3. Vacaciones 2024 / 2025 / 2026 ────────────────────────
    print("🏖️ Migrando vacaciones...")
    emp_cache = {}
    cur.execute("SELECT id, name FROM employees")
    for eid, ename in cur.fetchall():
        emp_cache[ename.lower()] = eid

    vacation_rows = []
    for sheet_year in [2024, 2025, 2026]:
        ws_vac = wb.get(str(sheet_year))
        if not ws_vac:
            continue
        for row in ws_vac.iter_rows(min_row=3, values_only=True):
            nombre = row[0]
            if not nombre or not isinstance(nombre, str):
                continue
            corr     = row[2]
            tomadas  = row[3]
            pend_ant = row[4]
            desc     = row[7]

            emp_id = emp_cache.get(str(nombre).strip().lower())
            if not emp_id:
                continue

            vacation_rows.append((
                sheet_year, emp_id,
                int(corr) if corr else None,
                int(tomadas) if tomadas else 0,
                int(pend_ant) if pend_ant else 0,
                str(desc) if desc and desc != 0 else None,
            ))

    if vacation_rows:
        execute_values(cur, """
            INSERT INTO vacation_records
              (year, employee_id, days_entitled, days_taken, pending_prev_year, description)
            VALUES %s
            ON CONFLICT (year, employee_id) DO NOTHING
        """, vacation_rows)
        print(f"  ✅ {len(vacation_rows)} registros de vacaciones migrados")

    conn.commit()
    cur.close()
    conn.close()
    print("\n✅ Migración completa.")


if __name__ == "__main__":
    migrate()
